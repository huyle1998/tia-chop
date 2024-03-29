from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.comments import Comment

def numOfItems(data):
    for i in range(1, len(data)):
        if str(data[i][0]).find("TỔNG CỘNG") >= 0:
            return i     

def forControl(A, B, row):
    for i in range(len(A)):
        if str(A[i]) == 'nan':
            A[i] = 0
        if str(B[i]) == 'nan':
            B[i] = 0

    if A[5] != B[5]:
        print("Ma don hang " + str(A[3]) + " bi KHAC SL dat don")
        ws['F'+str(row+7)].fill     = PatternFill("solid", fgColor="00FF0000")
        ws['F'+str(row+7)].comment  = Comment("LF: " + str(B[4]) + " \nTC: " + str(A[4]), 'Huy Le')
        ws['D'+str(row+7)].fill     = PatternFill("solid", fgColor="00FF0000")

    if A[9] != B[9]:
        print("Ma don hang " + str(A[3]) + " bi KHAC SL NCC cancel")
        ws['J'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['J'+str(row+7)].comment  = Comment("LF: " + str(B[6]) + " \nTC: " + str(A[6]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[10] != B[10]:
        print("Ma don hang " + str(A[3]) + " bi KHAC SL fail QC")
        ws['K'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['K'+str(row+7)].comment  = Comment("LF: " + str(B[7]) + " \nTC: " + str(A[7]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[6] != B[6]:
        print("Ma don hang " + str(A[3]) + " bi KHAC Don gia nhap kho")
        ws['G'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['G'+str(row+7)].comment  = Comment("LF: " + str(B[6]) + " \nTC: " + str(A[6]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[14] != B[14]:
        print("Ma don hang " + str(A[3]) + " bi KHAC Thanh tien nhap kho")
        ws['O'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['O'+str(row+7)].comment  = Comment("LF: " + str(B[14]) + " \nTC: " + str(A[14]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[15] != B[15]:
        print("Ma don hang " + str(A[3]) + " bi KHAC VAT")
        ws['P'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['P'+str(row+7)].comment  = Comment("LF: " + str(B[15]) + " \nTC: " + str(A[15]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[16] != B[16]:
        print("Ma don hang " + str(A[3]) + " bi KHAC Thanh tien thanh toan")
        ws['Q'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['Q'+str(row+7)].comment  = Comment("LF: " + str(B[16]) + " \nTC: " + str(A[16]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[17] != B[17]:
        print("Ma don hang " + str(A[3]) + " bi khac thanh tien can tru")
        ws['R'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['R'+str(row+7)].comment  = Comment("LF: " + str(B[17]) + " \nTC: " + str(A[17]), 'Huy Le')
        ws['D'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")



def main():
    global wb, ws

    shutil.copy2(os.getcwd() + '\TC.xlsx', os.getcwd() + '\ket_qua.xlsx')
    TC_data = pd.read_excel(os.getcwd() + '\TC.xlsx', header=5).values
    LF_data = pd.read_excel(os.getcwd() + '\LF.xlsx', header=5).values

    wb = load_workbook(os.getcwd() + '\ket_qua.xlsx')
    ws = wb[wb.sheetnames[0]]

    num_TC_item = numOfItems(TC_data)    
    num_LF_item = numOfItems(LF_data)    

    # Clean data
    TC_data = TC_data[0:num_TC_item]
    LF_data = LF_data[0:num_LF_item]

    # Get SKU and PO to compare
    is_TC_reserve = True
    LF_not_reserves = []
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O','P', 'Q', 'R']   
    for i in range(num_TC_item):
        for j in range(num_LF_item):
            if TC_data[i][2] == LF_data[j][2] and TC_data[i][3] == LF_data[j][3]:      # Consider updating new approximate comparation
                LF_not_reserves.append(j)
                forControl(TC_data[i][:], LF_data[j][:], i)
                is_TC_reserve = False
                break
            else:
                is_TC_reserve = True
        if True == is_TC_reserve:
            # TC_reserves.append(TC_data[i])
            for k in range(len(cols)):
                ws[cols[k] + str(i+7)]        = TC_data[i][k]
                ws[cols[k] + str(i+7)].fill   = PatternFill("solid", fgColor="09EA69")
    LF_reserves = []
    for j in range(num_LF_item):
        is_LF_reserve = True
        for LF_not_reserve in LF_not_reserves:
            if j == LF_not_reserve:
                is_LF_reserve = False
                break
            else:
                is_LF_reserve = True
        if is_LF_reserve:
            LF_reserves.append(LF_data[j])
    
    ws.insert_rows(num_TC_item+7, len(LF_reserves))   
    
    cnt_insert = 0
    for LF_reserve in LF_reserves:
        for i in range(len(cols)):
            ws[cols[i] + str(num_TC_item+7+cnt_insert)]        = LF_reserve[i]
            ws[cols[i] + str(num_TC_item+7+cnt_insert)].fill   = PatternFill("solid", fgColor="dfff3d")
        cnt_insert = cnt_insert + 1

    wb.save(os.getcwd() + '\ket_qua.xlsx')
    print("Hello users of Ngoc-Bui, this program is for you... enjoy it :)))")

if __name__ == "__main__":
    main()